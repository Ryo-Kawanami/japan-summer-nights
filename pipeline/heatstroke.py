"""熱中症による救急搬送者数を消防庁から取得する。

衛星の温度も気温も「何度だった」しか言わない。実際に人が倒れた数は、
その暑さが人にとって何を意味したかを直接示す唯一のデータになる。

出典: 総務省消防庁「熱中症による救急搬送状況」
      https://www.fdma.go.jp/disaster/heatstroke/post4.html

## データの形
  年ごとに xlsx 1ファイル。シートが月別（例 2024_05 … 2024_09）で、
  中身は 日付 × 都道府県コード の日次表。列に搬送人員（計）と年齢・重症度の内訳。

## 注意した点
  - 列の並びが年によって違う。2008年版は8列目が「年齢区分：不明」、
    2024年版は「傷病程度：死亡」。必ず列名で引くこと
  - URL のディレクトリが item/ と items/ の2種類ある。両方試す
  - 2008〜2010年は7月または6月からの集計で、5月分が無い年がある。
    夏（6〜8月）だけを使うぶんには影響しないが、年計を出すときは効く
  - 2008年からしかない。衛星データの2000年には遡れない
  - 搬送者数は人口と報告体制にも左右される。気温だけの関数ではない
"""

import pathlib
import re
import urllib.error
import urllib.request

import numpy as np

N_PREF = 47
YEARS = range(2008, 2025)          # 2025年は夏が終わっておらず確定値でない
SUMMER_MONTHS = (6, 7, 8)

BASE = "https://www.fdma.go.jp/disaster/heatstroke"
CACHE = pathlib.Path(__file__).resolve().parent / "cache" / "heatstroke"

COL_DATE = "日付"
COL_PREF = "都道府県コード"
COL_TOTAL = "搬送人員（計）"


def _era_name(year):
    """和暦のファイル名を作る。2018年までが平成、2019年からが令和。"""
    return f"h{year - 1988}" if year <= 2018 else f"r{year - 2018}"


def xlsx_path(year):
    return CACHE / f"{year}.xlsx"


def download(year, verbose=True):
    p = xlsx_path(year)
    if p.exists() and p.stat().st_size > 10000:
        return p
    p.parent.mkdir(parents=True, exist_ok=True)
    name = f"heatstroke003_data_{_era_name(year)}.xlsx"
    last = None
    for directory in ("items", "item"):
        url = f"{BASE}/{directory}/{name}"
        try:
            with urllib.request.urlopen(url, timeout=120) as r:
                body = r.read()
            if len(body) < 10000:          # 404 の HTML を掴んでいる
                continue
            p.write_bytes(body)
            if verbose:
                print(f"  {year}: {len(body)/1024:.0f}KB  {url}", flush=True)
            return p
        except urllib.error.HTTPError as e:
            last = e
    raise RuntimeError(f"{year} のファイルを取得できない: {last}")


def parse_year(year, verbose=True):
    """夏（6〜8月）の都道府県別搬送者数を長さ48の配列で返す。index 0 は未使用。"""
    import openpyxl

    wb = openpyxl.load_workbook(download(year, verbose), read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)      # 閉じた後にも使うので控えておく
    totals = np.zeros(N_PREF + 1, dtype="int64")
    days = 0
    try:
        matched = 0
        for sheet in sheet_names:
            # 区切り文字が年によって違う。2013年版は 2013_06 と 2013.07 が混在する。
            # ここを "_" だけで書くと7月と8月を黙って取りこぼす（実際に一度やった）。
            m = re.match(r"(\d{4})[._\-/](\d{1,2})$", sheet.strip())
            if not m:
                continue
            if int(m.group(2)) not in SUMMER_MONTHS:
                continue
            matched += 1
            ws = wb[sheet]
            rows = ws.iter_rows(values_only=True)
            header = [str(c).strip() if c is not None else "" for c in next(rows)]
            # 列の並びは年によって変わる。位置ではなく名前で引く。
            try:
                i_pref = header.index(COL_PREF)
                i_total = header.index(COL_TOTAL)
            except ValueError as e:
                raise RuntimeError(f"{year} {sheet}: 想定した列が無い {header[:5]}") from e

            seen_dates = set()
            for row in rows:
                if row[i_pref] is None or row[i_total] is None:
                    continue
                code = int(row[i_pref])
                if 1 <= code <= N_PREF:
                    totals[code] += int(row[i_total])
                seen_dates.add(row[header.index(COL_DATE)])
            days += len(seen_dates)
    finally:
        wb.close()

    # 6・7・8月の3シートが揃わない年は、区切り文字違いなどで取りこぼしている。
    # 黙って少ない数字を返すのが一番まずいので、ここで落とす。
    if matched != len(SUMMER_MONTHS):
        raise RuntimeError(
            f"{year}: 夏のシートが {matched}/{len(SUMMER_MONTHS)} しか見つからない。"
            f"シート名 {sheet_names}")
    if verbose:
        print(f"  {year}: 夏{days}日  全国 {totals.sum():,}人", flush=True)
    return totals


def load():
    """(years, counts) を返す。counts は (年数, 48) で index 0 は未使用。"""
    npz = CACHE / "summer_by_prefecture.npz"
    if npz.exists():
        z = np.load(npz)
        return z["years"], z["counts"]

    years, rows = [], []
    for y in YEARS:
        try:
            rows.append(parse_year(y))
            years.append(y)
        except Exception as e:
            print(f"  失敗 {y}: {type(e).__name__}: {e}", flush=True)
    years = np.array(years)
    counts = np.stack(rows)
    CACHE.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(npz, years=years, counts=counts)
    return years, counts


if __name__ == "__main__":
    years, counts = load()
    print(f"\n{len(years)}年ぶん（{years[0]}〜{years[-1]}）")
    nat = counts[:, 1:].sum(axis=1)
    for y, n in zip(years, nat):
        print(f"  {y}: {n:7,}人")
